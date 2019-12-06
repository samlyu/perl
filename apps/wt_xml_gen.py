#*****************************************************************************************
#-----------------------------------------------------------------------------------------
# Last change date : $Date: 2019.11.25$
# Revision         : $Revision: 1.2$
# Function         : Generate XML reglist from Excel spreadsheet data
#-----------------------------------------------------------------------------------------
#******************************************************************************************

import  re,os,sys,math
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import shutil
from pathlib import Path
from collections import defaultdict
from ast import literal_eval

print('Python Script for Reglist Generation')
xls_name = input('Input MS Excel file name: ')
wb = load_workbook(xls_name + '.xlsx')
sheet_name_list = wb.sheetnames
all_reg_msg_dict = defaultdict(list)
reg_dict = dict()
print('Register address list: ')
for t in range(len(sheet_name_list)):
    module_name = sheet_name_list[t]
    sheet_group = wb[module_name]
    param_value = list()
    for row_data2 in sheet_group.iter_rows(min_row=2, max_row=2):
        for cell2 in row_data2:
            if cell2.value is not None:
                param_value.append(cell2.value)
    address_width = param_value[0]
    base_address = hex(literal_eval(re.sub(r'.*\'h', '0x', param_value[1])))
    data_width = param_value[2]
    offset_address_width = param_value[3]

    m_list=list()
    cr=list()
    row_list =list()
    m_list = sheet_group.merged_cells
    for cell in m_list:
        r1,r2,c1,c2 = cell.min_row,cell.max_row,cell.min_col,cell.max_col
        if r2 - r1 > 0:
            if c1 == 1:
                cr.append((r1,r2))
                for i in range (r1,r2+1):
                    row_list.append(i)
                    row_list= sorted(row_list)
                cr = sorted(cr)
    counter = 0
    all_reg_msg = list()
    flag = 0
    num = 0
    for row in range(4,sheet_group.max_row+1):
        if row in row_list :
            num0 = cr[counter][0]
            num1 = cr[counter][1]
            num = num1 - num0
            if row == num0:
                flag = 1
            elif row > num0 and row < num1:
                flag = 2
            elif row == num1:
                flag = 3
                counter = counter + 1
            elif row == num1 + 1:
                flag = 0
            all_reg_msg.append((row, flag, num))
        else :
            flag = 0
            num = 0
            all_reg_msg.append((row, flag, num))

    for i in range(len(all_reg_msg)):
        reg_row  = all_reg_msg[i][0]
        reg_flag = all_reg_msg[i][1]
        reg_num  = all_reg_msg[i][2]
        if reg_flag == 0:
            reg_addr_offset = hex(literal_eval(sheet_group['A'+str(reg_row)].value))
            reg_addr = hex(int(reg_addr_offset, 0) + int(base_address, 0))
            print(reg_addr)
            reg_name = sheet_group['B'+str(reg_row)].value
            reg_value = hex(literal_eval(re.sub(r'.*\'h', '0x', sheet_group['C'+str(reg_row)].value)))
            bit_reset = sheet_group['D'+str(reg_row)].value
            if bit_reset is not None:
                bit_flag = re.findall(r'.*\'b', bit_reset)
                if bit_flag:
                    bit_reset = hex(int(re.sub(r'.*\'b', '0b', bit_reset), 2))
                else:
                    bit_flag_h = re.findall(r'.*\'h', bit_reset)
                    if bit_flag_h:
                        bit_reset = hex(literal_eval(re.sub(r'.*\'h', '0x', bit_reset)))
                    else:
                        bit_flag_d = re.findall(r'.*\'d', bit_reset)
                        if bit_flag_d:
                            bit_reset = hex(int(re.sub(r'.*\'d', '', bit_reset)))
            else:
                print('\tReg has uncertain bit reset value')
            bit_range = sheet_group['E'+str(reg_row)].value
            bit_name = sheet_group['F'+str(reg_row)].value
            bit_attr = sheet_group['G'+str(reg_row)].value
            bit_cons = sheet_group['I'+str(reg_row)].value
            if bit_cons is not None:
                bit_cons = bit_cons.replace('\n', ' ')
            bit_desc = sheet_group['J'+str(reg_row)].value
            if bit_desc is not None:
                bit_desc = bit_desc.replace('\n', ' ')
            reg_size = 0
            if bit_range is not None:
                data_range = re.findall(r'\[(\d+)\:(\d+)\]', bit_range)
                if data_range:
                    pos_msb = data_range[0][0]
                    pos_lsb = data_range[0][1]
                else:
                    data_range = re.findall(r'\[(\d+)\]', bit_range)
                    pos_msb = data_range[0][0]
                    pos_lsb = data_range[0][0]
            else:
                data_range = 'None'
                print('\tReg has empty bit field: ')
                print('\t\treg_name: ', reg_name)
                print('\t\tdata_range: ', data_range)
            reg_msg =[reg_addr, reg_name, reg_value, bit_name, bit_reset, bit_range, bit_attr, \
                pos_lsb, pos_msb, module_name, bit_cons, bit_desc, \
                address_width, base_address, data_width, offset_address_width, reg_addr_offset, reg_size]
            all_reg_msg_dict[reg_addr].append(reg_msg)
            reg_dict.update({reg_addr: module_name})

        elif reg_flag == 1:
            reg_addr_offset = hex(literal_eval(sheet_group['A'+str(reg_row)].value))
            reg_addr = hex(int(reg_addr_offset, 0) + int(base_address, 0))
            print(reg_addr)
            reg_name = sheet_group['B' + str(reg_row)].value
            reg_value = hex(literal_eval(re.sub(r'.*\'h', '0x', sheet_group['C' + str(reg_row)].value)))
            for i in range(0, reg_num+1):
                bit_reset = sheet_group['D' + str(reg_row+i)].value
                if bit_reset is not None:
                    bit_flag = re.findall(r'.*\'b', bit_reset)
                    if bit_flag:
                        bit_reset = hex(int(re.sub(r'.*\'b', '0b', bit_reset), 2))
                    else:
                        bit_flag_h = re.findall(r'.*\'h', bit_reset)
                        if bit_flag_h:
                            bit_reset = hex(literal_eval(re.sub(r'.*\'h', '0x', bit_reset)))
                        else:
                            bit_flag_d = re.findall(r'.*\'d', bit_reset)
                            if bit_flag_d:
                                bit_reset = hex(int(re.sub(r'.*\'d', '', bit_reset)))
                else:
                    print('\tReg has uncertain bit reset value')
                bit_range = sheet_group['E' + str(reg_row+i)].value
                bit_name = sheet_group['F' + str(reg_row+i)].value
                bit_attr = sheet_group['G' + str(reg_row+i)].value
                bit_cons = sheet_group['I'+str(reg_row+i)].value
                if bit_cons is not None:
                    bit_cons = bit_cons.replace('\n', ' ')
                bit_desc = sheet_group['J'+str(reg_row+i)].value
                if bit_desc is not None:
                    bit_desc = bit_desc.replace('\n', ' ')
                reg_size = 0
                if bit_range is not None:
                    data_range1 = re.findall(r'\[(\d+)\:(\d+)\]', bit_range)
                    data_range2 = re.findall(r'\[(\d+)\]', bit_range)
                    if data_range1:
                        pos_msb = data_range1[0][0]
                        pos_lsb = data_range1[0][1]
                    elif data_range2:
                        pos_msb = data_range2[0][0]
                        pos_lsb = data_range2[0][0]
                    else:
                        print('\tReg has uncertain bit range: ')
                        print('\t\treg_name: ', reg_name)
                        print('\t\tbit name: ', bit_name)
                        pos_msb = 0;
                        pos_lsb = 0;
                else:
                    data_range = 'None'
                    print('\tReg has empty bit field: ')
                    print('\t\treg_name: ', reg_name)
                    print('\t\tdata_range: ', data_range)

                reg_msg = [reg_addr, reg_name, reg_value, bit_name, bit_reset, bit_range, bit_attr, \
                    pos_lsb, pos_msb, module_name, bit_cons, bit_desc, \
                    address_width, base_address, data_width, offset_address_width, reg_addr_offset, reg_size]
                all_reg_msg_dict[reg_addr].append(reg_msg)
                reg_dict.update({reg_addr: module_name})

reg_addr_list = list(all_reg_msg_dict.keys())

temp_list = list()
for y in reg_dict.values():
    temp_list.append(y)
temp_set = set(temp_list)
mod_dict = {}
for x in temp_set:
    mod_dict.update({x: temp_list.count(x)})
for x, y in all_reg_msg_dict.items():
    for z in mod_dict.keys():
        if z == y[0][9]:
            y[0][17] = mod_dict[z]

# print(all_reg_msg_dict)

SOC_DNN = open('SOC_DNN.xml','w')
SOC_DNN.truncate()
SOC_DNN.write('<?xml version=\"1.0\" encoding=\"utf-8\" standalone=\"no\"?>\n')
SOC_DNN.write('<device schemaVersion=\"1.1\"\n')
SOC_DNN.write('xmlns:xs=\"http://www.w3.org/2001/XMLSchema-instance\"\n')
SOC_DNN.write('xs:noNamespaceSchemaLocation=\"CMSIS-SVD_Schema_1_1.xsd\">\n')
SOC_DNN.write('\t<name>SOC_DNN</name>\n')
SOC_DNN.write('\t<version>1.1</version>\n')
SOC_DNN.write('\t<description>SOC_DNN</description>\n')
SOC_DNN.write('\t<peripherals>\n')

module_iter = 0
module_current = 'None'

for t in range(len(reg_addr_list)):
    SOC_reg_addr = reg_addr_list[t]
    if type(all_reg_msg_dict[SOC_reg_addr]) == list :
        SOC_reg_msg = all_reg_msg_dict[SOC_reg_addr]

    for z in range(len(SOC_reg_msg)):
        SOC_reg_name = SOC_reg_msg[z][1]
        SOC_reg_reset = SOC_reg_msg[z][2]
        SOC_field_name = SOC_reg_msg[z][3]
        SOC_field_reset = SOC_reg_msg[z][4]
        SOC_field_attr = SOC_reg_msg[z][6]
        SOC_field_pos_lsb = SOC_reg_msg[z][7]
        SOC_field_pos_msb = SOC_reg_msg[z][8]        
        SOC_module_name = SOC_reg_msg[z][9]
        SOC_field_cons = SOC_reg_msg[z][10]
        SOC_field_desc = SOC_reg_msg[z][11]
        SOC_address_width = SOC_reg_msg[z][12]
        SOC_base_address = SOC_reg_msg[z][13]
        SOC_data_width = SOC_reg_msg[z][14]
        SOC_offset_address_width = SOC_reg_msg[z][15]
        SOC_reg_addr_offset = SOC_reg_msg[z][16]
        SOC_reg_size = SOC_reg_msg[z][17]

        if SOC_module_name != module_current:
            module_current = SOC_module_name
            SOC_DNN.write('\t\t<peripheral>\n')
            SOC_DNN.write('\t\t\t<name>' + SOC_module_name + '</name>\n')
            SOC_DNN.write('\t\t\t<description>' + SOC_module_name + ' registers</description>\n')
            SOC_DNN.write('\t\t\t<groupName>' + SOC_module_name + '</groupName>\n')
            SOC_DNN.write('\t\t\t<baseAddress>' + SOC_base_address + '</baseAddress>\n')
            SOC_DNN.write('\t\t\t<addressBlock>\n\t\t\t\t<offset>' + SOC_reg_addr_offset + '</offset>\n')
            SOC_DNN.write('\t\t\t\t<size>' + hex(SOC_reg_size*SOC_data_width) + '</size>\n')
            SOC_DNN.write('\t\t\t\t<usage>registers</usage>\n')
            SOC_DNN.write('\t\t\t</addressBlock>\n')
            SOC_DNN.write('\t\t\t<registers>\n')

        if z == 0:
            SOC_DNN.write('\t\t\t\t<register>\n')
            SOC_DNN.write('\t\t\t\t\t<name>' + SOC_reg_name + '</name>\n')
            SOC_DNN.write('\t\t\t\t\t<displayName>' + SOC_reg_name + '</displayName>\n')
            SOC_DNN.write('\t\t\t\t\t<description>' + SOC_reg_name + '</description>\n')
            SOC_DNN.write('\t\t\t\t\t<addressOffset>' + SOC_reg_addr + '</addressOffset>\n')
            SOC_DNN.write('\t\t\t\t\t<size>' + hex(SOC_data_width) + '</size> \n')
            SOC_DNN.write('\t\t\t\t\t<access>' + str(SOC_field_attr) + '</access> \n')
            SOC_DNN.write('\t\t\t\t\t<resetValue>' + SOC_reg_reset + '</resetValue>\n')
            SOC_DNN.write('\t\t\t\t\t<fields>\n')

        SOC_DNN.write('\t\t\t\t\t\t<field> \n')
        SOC_DNN.write('\t\t\t\t\t\t\t<name>' + str(SOC_field_name) + '</name>\n')
        SOC_DNN.write('\t\t\t\t\t\t\t<description>' + str(SOC_field_desc) + '</description>\n')
        SOC_DNN.write('\t\t\t\t\t\t\t<constraint>' + str(SOC_field_cons) + '</constraint>\n')
        SOC_DNN.write('\t\t\t\t\t\t\t<bitOffset>'+ hex(int(SOC_field_pos_lsb)) + '</bitOffset>\n')
        bitWidth = int(SOC_field_pos_msb) - int(SOC_field_pos_lsb) + 1
        SOC_DNN.write('\t\t\t\t\t\t\t<bitWidth>' + hex(bitWidth) +'</bitWidth>\n')
        SOC_DNN.write('\t\t\t\t\t\t\t<bitValue>' + str(SOC_field_reset) + '</bitValue>\n')
        SOC_DNN.write('\t\t\t\t\t\t</field>\n')

        if z == len(SOC_reg_msg) - 1:
            SOC_DNN.write('\t\t\t\t\t</fields>\n')
            SOC_DNN.write('\t\t\t\t</register>\n')

        if t != len(reg_addr_list)-1:
            SOC_reg_addr_next = reg_addr_list[t+1]
        if type(all_reg_msg_dict[SOC_reg_addr_next]) == list :
            SOC_reg_msg_next = all_reg_msg_dict[SOC_reg_addr_next]
        for zz in range(len(SOC_reg_msg_next)):
            SOC_module_name_next = SOC_reg_msg_next[zz][9]
        if SOC_module_name_next != module_current and (z == len(SOC_reg_msg) - 1):
            SOC_DNN.write('\t\t\t</registers>\n')
            SOC_DNN.write('\t\t</peripheral>\n')

SOC_DNN.write('\t\t\t<registers>\n')
SOC_DNN.write('\t\t</peripheral>\n')
SOC_DNN.write('\t</peripherals>\n')
SOC_DNN.write('</device>\n')
SOC_DNN.close()

print('XML Generated')

